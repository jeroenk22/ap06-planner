"""Aanvullende tests voor xlsx_parser — workbook fixtures en edge cases."""

from datetime import datetime
from io import BytesIO

import pytest
from openpyxl import Workbook

from ap06_planner.models.schemas import PlanningRegel
from ap06_planner.parsers.xlsx_parser import (
    _cel,
    _datum_uit_tabnaam,
    _locatie_tekst,
    detecteer_datum,
    detecteer_dagnaam,
    detecteer_headers,
    is_eurofins_formaat,
    lees_planningsbestand,
    selecteer_tabbladen,
)


def _maak_workbook_bytes(tab_naam="13-4 Maandag", met_data=True) -> BytesIO:
    """Maak een minimaal geldig AP06 planningsbestand in-memory."""
    wb = Workbook()
    ws = wb.active
    ws.title = tab_naam

    # Rij 1: datum in kolom C (index 2, kolom C)
    ws.cell(row=1, column=3, value=datetime(2026, 4, 13))

    # Rij 3: headers (standaard formaat)
    ws.cell(row=3, column=2, value="Monsternemer")   # B
    ws.cell(row=3, column=4, value="Wijzigingen")    # D
    ws.cell(row=3, column=5, value="Locatie")        # E
    ws.cell(row=3, column=6, value="Klant")          # F

    if met_data:
        # Rij 4: dataregel
        ws.cell(row=4, column=2, value="Jan de Vries")
        ws.cell(row=4, column=5, value="Bladel TonTrans 7-18 LAD17")

        # Rij 5: overgeslagen regel
        ws.cell(row=5, column=2, value="Piet Smit")
        ws.cell(row=5, column=4, value="vervallen")
        ws.cell(row=5, column=5, value="Zundert Dams 6-18 LAD1")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _maak_eurofins_bytes() -> BytesIO:
    """Maak een Eurofins Agro formaat planningsbestand."""
    wb = Workbook()
    ws = wb.active
    ws.title = "13-4 Maandag"

    # Datum in kolom B (Eurofins formaat, index 1)
    ws.cell(row=1, column=2, value=datetime(2026, 4, 13))

    # Headers op rij 2: kolom A=Monsternemer, D=wijzigingen, E=Klant, B=Laadlocatie
    ws.cell(row=2, column=1, value="Monsternemer")
    ws.cell(row=2, column=2, value="Laadlocatie")
    ws.cell(row=2, column=4, value="Wijzigingen")
    ws.cell(row=2, column=5, value="Klant")

    ws.cell(row=3, column=1, value="Anouk Bakker")
    ws.cell(row=3, column=5, value="Marrum JB 7-9 LOS")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


class TestLocatieTekst:
    def test_locatie_met_tijdvenster(self):
        result = _locatie_tekst("Bladel 7-18 LAD17", None)
        assert result == "Bladel 7-18 LAD17"

    def test_klant_met_tijdvenster(self):
        result = _locatie_tekst(None, "Marrum JB 7-9 LOS")
        assert result == "Marrum JB 7-9 LOS"

    def test_beide_met_tijdvenster(self):
        # Alleen "Bladel 7-18" heeft een tijdpatroon, "LAD17" niet
        result = _locatie_tekst("Bladel 7-18", "LAD17")
        assert result == "Bladel 7-18"

    def test_beide_met_tijdvenster_in_klant(self):
        # Beide hebben tijdpatroon → combineer
        result = _locatie_tekst("Bladel 7-18", "Zundert 6-9")
        assert result == "Bladel 7-18 Zundert 6-9"

    def test_geen_tijdvenster_combineer(self):
        result = _locatie_tekst("Bladel", "Klant")
        assert result == "Bladel Klant"

    def test_alleen_locatie_geen_klant(self):
        result = _locatie_tekst("Bladel", None)
        assert result == "Bladel"

    def test_alleen_klant_geen_locatie(self):
        result = _locatie_tekst(None, "Klant")
        assert result == "Klant"

    def test_beide_none(self):
        result = _locatie_tekst(None, None)
        assert result is None


class TestCel:
    def test_normale_waarde(self):
        row = ("a", "b", "c")
        assert _cel(row, 1) == "b"

    def test_whitespace_gestript(self):
        row = ("  spaties  ", "")
        assert _cel(row, 0) == "spaties"

    def test_lege_string_geeft_none(self):
        row = ("", "a")
        assert _cel(row, 0) is None

    def test_none_waarde_geeft_none(self):
        row = (None, "b")
        assert _cel(row, 0) is None

    def test_index_buiten_bereik(self):
        row = ("a",)
        assert _cel(row, 5) is None

    def test_none_index(self):
        row = ("a", "b")
        assert _cel(row, None) is None

    def test_getal_wordt_string(self):
        row = (42, "b")
        assert _cel(row, 0) == "42"


class TestDatumUitTabnaam:
    def test_normaal(self):
        result = _datum_uit_tabnaam("13-4 Maandag")
        assert result is not None
        assert "13" in result
        assert "04" in result

    def test_ongeldig_patroon(self):
        assert _datum_uit_tabnaam("Blad2") is None

    def test_ongeldige_datum(self):
        # 32-13 is ongeldig
        result = _datum_uit_tabnaam("32-13 Maandag")
        assert result is None


class TestDetecteerDatum:
    def test_standaard_formaat(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=3, value=datetime(2026, 4, 13))
        result = detecteer_datum(ws, eurofins_formaat=False)
        assert result == "13-04-2026"

    def test_eurofins_formaat(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=2, value=datetime(2026, 4, 13))
        result = detecteer_datum(ws, eurofins_formaat=True)
        assert result == "13-04-2026"

    def test_geen_datum_geeft_none(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=3, value="geen datum")
        result = detecteer_datum(ws, eurofins_formaat=False)
        assert result is None


class TestDetecteerDagnaam:
    def test_vindt_maandag(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="maandag")
        assert detecteer_dagnaam(ws) == "maandag"

    def test_vindt_case_insensitive(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Vrijdag")
        assert detecteer_dagnaam(ws) == "vrijdag"

    def test_geen_dagnaam(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Niet een dag")
        assert detecteer_dagnaam(ws) is None


class TestDetecteerHeaders:
    def test_vindt_headers_op_rij_3(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=3, column=2, value="Monsternemer")
        ws.cell(row=3, column=4, value="Wijzigingen")
        rij_nr, kolommap = detecteer_headers(ws)
        assert rij_nr == 3
        assert "monsternemer" in kolommap
        assert "wijzigingen" in kolommap

    def test_geen_headers(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Iets anders")
        rij_nr, kolommap = detecteer_headers(ws)
        assert rij_nr is None
        assert kolommap is None

    def test_eurofins_headers_op_rij_2(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=2, column=1, value="Monsternemer")
        ws.cell(row=2, column=2, value="Laadlocatie")
        rij_nr, kolommap = detecteer_headers(ws)
        assert rij_nr == 2
        assert "laadlocatie" in kolommap


class TestIsEurofinsFOrmaat:
    def test_eurofins_herkend(self):
        assert is_eurofins_formaat({"laadlocatie": 1, "monsternemer": 0}) is True

    def test_standaard_formaat(self):
        assert is_eurofins_formaat({"locatie": 4, "monsternemer": 1}) is False


class TestLeesPlanningsbestand:
    def test_standaard_bestand(self):
        buf = _maak_workbook_bytes()
        resultaten = lees_planningsbestand(buf)
        assert len(resultaten) == 1
        tab = resultaten[0]
        assert tab["tabblad"] == "13-4 Maandag"
        assert tab["datum"] == "13-04-2026"
        assert "maandag" in (tab["dagnaam"] or "")
        regels = tab["regels"]
        niet_overgeslagen = [r for r in regels if not r.overgeslagen]
        overgeslagen = [r for r in regels if r.overgeslagen]
        assert len(niet_overgeslagen) == 1
        assert niet_overgeslagen[0].monsternemer_naam == "Jan de Vries"
        assert len(overgeslagen) == 1  # Piet Smit → vervallen

    def test_geen_geldige_tabs_gebruikt_eerste(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Overzicht"  # Geen datum-dagnaam patroon
        ws.cell(row=3, column=2, value="Monsternemer")
        ws.cell(row=4, column=2, value="Test Persoon")
        ws.cell(row=4, column=5, value="Bladel 7-18 LAD")
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resultaten = lees_planningsbestand(buf)
        assert len(resultaten) == 1

    def test_corrupt_bestand_raises(self):
        with pytest.raises(ValueError, match="Kan planningsbestand"):
            lees_planningsbestand(BytesIO(b"geen xlsx data"))

    def test_tab_zonder_headers_overgeslagen(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "13-4 Maandag"
        ws.cell(row=1, column=1, value="Geen header hier")
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resultaten = lees_planningsbestand(buf)
        # Tab zonder 'Monsternemer' header wordt overgeslagen
        assert len(resultaten) == 0

    def test_meerdere_tabs(self):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "13-4 Maandag"
        ws1.cell(row=3, column=2, value="Monsternemer")
        ws1.cell(row=4, column=2, value="Jan de Vries")
        ws1.cell(row=4, column=5, value="Bladel 7-18 LAD")

        ws2 = wb.create_sheet("14-4 Dinsdag")
        ws2.cell(row=3, column=2, value="Monsternemer")
        ws2.cell(row=4, column=2, value="Piet Smit")
        ws2.cell(row=4, column=5, value="Zundert 6-18 LAD")

        wb.create_sheet("Blad2")  # Utility tab → overgeslagen

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resultaten = lees_planningsbestand(buf)
        assert len(resultaten) == 2

    def test_lege_rijen_overgeslagen(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "13-4 Maandag"
        ws.cell(row=3, column=2, value="Monsternemer")
        ws.cell(row=3, column=5, value="Locatie")
        # Rij 4: volledig leeg
        # Rij 5: geen monsternemer
        ws.cell(row=5, column=5, value="Bladel 7-18 LAD")
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resultaten = lees_planningsbestand(buf)
        assert len(resultaten[0]["regels"]) == 0

    def test_eurofins_formaat(self):
        buf = _maak_eurofins_bytes()
        resultaten = lees_planningsbestand(buf)
        assert len(resultaten) == 1
        regels = resultaten[0]["regels"]
        assert len(regels) == 1
        assert regels[0].monsternemer_naam == "Anouk Bakker"

    def test_dagnaam_uit_tabnaam(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "13-4 Vrijdag"
        ws.cell(row=3, column=2, value="Monsternemer")
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)
        resultaten = lees_planningsbestand(buf)
        assert resultaten[0]["dagnaam"] == "vrijdag"
