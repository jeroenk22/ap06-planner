"""
xlsx_parser.py — Leest AP06-planningsbestanden in.

Kritieke bevindingen uit analyse van 6 voorbeeldbestanden:
- Laad ALTIJD met data_only=True (anders formules zoals =AK4 in plaats van waarden)
- Tabblad-selectie: kies tabs met "dd-mm Dagnaam" patroon, sla Blad2/Blad129 over
- Header-detectie: scan rijen 1-5 op aanwezigheid van 'Monsternemer'
- 2 kolomformaten: standaard (E=Locatie, F=Klant) en Eurofins (E=Klant)
- Datum: rij 1 kolom C (of B bij Eurofins) als datetime object
- Leeg monsternemer = skip; "vervallen"/"intrekken"/"ingetrokken" in wijzigingen = skip
"""

import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from ap06_planner.models.schemas import PlanningRegel

# Patroon voor geldige plannings-tabbladen: "11-5 maandag", "13-4 Maandag", "9-5 Zaterdag"
GELDIG_TABBLAD_PATROON = re.compile(
    r"\d{1,2}-\d{1,2}\s+(maandag|dinsdag|woensdag|donderdag|vrijdag|zaterdag|zondag)",
    re.IGNORECASE,
)

_TIJDPATROON = re.compile(r"\b\d{1,2}(?:[.,]\d{1,2})?\s*-\s*\d{1,2}(?:[.,]\d{1,2})?\b")


def _locatie_tekst(locatie: str | None, klant: str | None) -> str | None:
    """Kies de cel die het tijdvenster bevat, ongeacht kolomnaam."""
    loc_heeft_tijd = bool(locatie and _TIJDPATROON.search(locatie))
    klt_heeft_tijd = bool(klant and _TIJDPATROON.search(klant))
    if loc_heeft_tijd and klt_heeft_tijd:
        return f"{locatie} {klant}"
    if loc_heeft_tijd:
        return locatie
    if klt_heeft_tijd:
        return klant
    if locatie and klant:
        return f"{locatie} {klant}"
    return locatie or klant or None


SKIP_WIJZIGINGEN = re.compile(
    r"\b(vervallen|intrekken|ingetrokken)\b",
    re.IGNORECASE,
)


def selecteer_tabbladen(sheetnames: list[str]) -> list[str]:
    """Selecteer tabbladen met een datumDagnaam-patroon. Sla utility-tabs over."""
    return [s for s in sheetnames if GELDIG_TABBLAD_PATROON.search(s)]


def detecteer_datum(ws, eurofins_formaat: bool) -> str | None:
    """Extraheer de planningsdatum (dd-mm-jjjj) uit de eerste rij."""
    col_idx = 1 if eurofins_formaat else 2  # B of C (0-indexed)
    for row in ws.iter_rows(min_row=1, max_row=3, values_only=True):
        if len(row) > col_idx and isinstance(row[col_idx], datetime):
            dt = row[col_idx]
            return dt.strftime("%d-%m-%Y")
    return None


def _datum_uit_tabnaam(tab_naam: str) -> str | None:
    """Fallback: haal datum uit tabbladnaam zoals '20-4 maandag'."""
    match = re.search(r"(\d{1,2})-(\d{1,2})", tab_naam)
    if not match:
        return None
    dag, maand = int(match.group(1)), int(match.group(2))
    jaar = date.today().year
    try:
        return datetime(jaar, maand, dag).strftime("%d-%m-%Y")
    except ValueError:
        return None


def detecteer_dagnaam(ws) -> str | None:
    """Extraheer de dagnaam (bijv. 'maandag') uit rij 1."""
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        for cel in row:
            if isinstance(cel, str) and cel.lower() in {
                "maandag",
                "dinsdag",
                "woensdag",
                "donderdag",
                "vrijdag",
                "zaterdag",
                "zondag",
            }:
                return cel.lower()
    return None


def detecteer_headers(ws) -> tuple[int, dict[str, int]] | tuple[None, None]:
    """
    Zoek de headerrij (max rij 5) en retourneer (rijnummer, kolom→index dict).
    Ondersteunt zowel standaard- als Eurofins Agro-formaat.
    """
    for rij_nr, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        row_lower = [str(v).strip().lower() if v is not None else "" for v in row]
        if "monsternemer" in row_lower:
            kolommap = {}
            for idx, header in enumerate(row_lower):
                if header:
                    kolommap[header] = idx
            return rij_nr, kolommap
    return None, None


def is_eurofins_formaat(kolommap: dict[str, int]) -> bool:
    """
    Eurofins Agro formaat heeft 'laadlocatie' als kolom, standaard heeft 'locatie'.
    In Eurofins staat het tijdvenster in de 'klant' kolom (niet in 'locatie').
    """
    return "laadlocatie" in kolommap


def lees_planningsbestand(
    bron: Path | BytesIO,
) -> list[dict]:
    """
    Hoofdfunctie: lees een xlsx-planningsbestand in.
    Retourneert een lijst van dicts met {tabblad, datum, dagnaam, regels: [PlanningRegel]}.
    """
    try:
        wb = load_workbook(bron, data_only=True, read_only=False)
    except Exception as e:
        raise ValueError(f"Kan planningsbestand niet laden: {e}") from e

    geselecteerde_tabs = selecteer_tabbladen(wb.sheetnames)
    if not geselecteerde_tabs:
        # Fallback: gebruik het eerste tabblad als er geen match is
        geselecteerde_tabs = [wb.sheetnames[0]]

    resultaten = []

    for tab_naam in geselecteerde_tabs:
        ws = wb[tab_naam]

        header_rij, kolommap = detecteer_headers(ws)
        if header_rij is None:
            continue

        eurofins = is_eurofins_formaat(kolommap)
        datum = detecteer_datum(ws, eurofins_formaat=eurofins) or _datum_uit_tabnaam(tab_naam)
        dagnaam = detecteer_dagnaam(ws)

        # Tabblad-naam geeft ook datuminfo: "13-4 Maandag"
        if dagnaam is None:
            match = GELDIG_TABBLAD_PATROON.search(tab_naam)
            if match:
                dagnaam = match.group(1).lower()

        regels = _verwerk_rijen(ws, header_rij, kolommap, eurofins)

        resultaten.append(
            {
                "tabblad": tab_naam,
                "datum": datum,
                "dagnaam": dagnaam,
                "regels": regels,
                "kolommap": dict(kolommap),
            }
        )

    wb.close()
    return resultaten


def _detecteer_tijdvenster_kolom(ws, header_rij: int, skip_kolommen: set[int]) -> int | None:
    """
    Scan de eerste datarijen en retourneer de kolomindex die het vaakst
    een tijdpatroon bevat (bijv. '7-18', '8.30-10.30'), ongeacht de kolomnaam.
    """
    scores: dict[int, int] = {}
    for row in ws.iter_rows(
        min_row=header_rij + 1,
        max_row=min(header_rij + 10, ws.max_row or header_rij + 10),
        values_only=True,
    ):
        for idx, cel in enumerate(row):
            if idx >= 10 or idx in skip_kolommen:  # alleen kolommen A t/m J
                continue
            if cel and isinstance(cel, str) and _TIJDPATROON.search(cel):
                scores[idx] = scores.get(idx, 0) + 1
    return max(scores, key=scores.get) if scores else None


def _verwerk_rijen(
    ws,
    header_rij: int,
    kolommap: dict[str, int],
    eurofins: bool,
) -> list[PlanningRegel]:
    """Verwerk alle datarijen onder de headerrij."""
    regels = []

    idx_monsternemer = kolommap.get("monsternemer")
    idx_wijzigingen = kolommap.get("wijzigingen")
    idx_locatie_naam = kolommap.get("locatie") or kolommap.get("laadlocatie")

    # Detecteer welke kolom tijdvenster-data bevat op basis van celinhoud
    skip = {i for i in [idx_monsternemer, idx_wijzigingen] if i is not None}
    idx_tijdvenster = _detecteer_tijdvenster_kolom(ws, header_rij, skip)

    for row in ws.iter_rows(min_row=header_rij + 1, values_only=True):
        # Skip volledig lege rijen
        if all(v is None for v in row):
            continue

        monsternemer_naam = _cel(row, idx_monsternemer)
        if not monsternemer_naam:
            continue  # Geen monsternemer = skip

        wijzigingen = _cel(row, idx_wijzigingen)
        tv_tekst = _cel(row, idx_tijdvenster) if idx_tijdvenster is not None else None
        loc_naam = (
            _cel(row, idx_locatie_naam)
            if (idx_locatie_naam is not None and idx_locatie_naam != idx_tijdvenster)
            else None
        )

        locatie_raw = _locatie_tekst(loc_naam, tv_tekst)
        klant_raw = None

        # Skip bij "vervallen" / "intrekken" / "ingetrokken"
        if wijzigingen and SKIP_WIJZIGINGEN.search(wijzigingen):
            regels.append(
                PlanningRegel(
                    monsternemer_naam=monsternemer_naam,
                    wijzigingen=wijzigingen,
                    locatie_raw=locatie_raw,
                    klant_raw=klant_raw,
                    overgeslagen=True,
                    reden_overgeslagen=f"wijzigingen: '{wijzigingen}'",
                )
            )
            continue

        regels.append(
            PlanningRegel(
                monsternemer_naam=monsternemer_naam,
                wijzigingen=wijzigingen,
                locatie_raw=locatie_raw,
                klant_raw=klant_raw,
            )
        )

    return regels


def _cel(row: tuple, idx: int | None) -> str | None:
    """Haal celwaarde op en strip whitespace. Geeft None bij lege cellen."""
    if idx is None or idx >= len(row):
        return None
    val = row[idx]
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None
