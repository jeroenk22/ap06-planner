"""
beheer.py — Beheer de monsternemer-database via Streamlit UI.

Functies:
- Overzicht van alle monsternemers
- Monsternemer toevoegen / bewerken
- Monsternemer verwijderen
- Importeren vanuit xlsx (AP06 mockdata formaat)
"""

import uuid
from io import BytesIO

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from ap06_planner.models.schemas import Monsternemer
from ap06_planner.services.db_service import (
    haal_alle_monsternemers,
    initialiseer_db,
    update_monsternemer,
    verwijder_monsternemer,
    voeg_monsternemer_toe,
)

_UITERLIJKE_TIJD_HELP = (
    "Laatste tijdstip waarop de monsternemer opgehaald wil worden. "
    "Wordt getoond als gewensttijd — leidt NIET automatisch tot doorschuiven."
)
_UITERLIJKE_PLANTIJD_HELP = (
    "Planningtechnische grens: als de berekende aankomsttijd hier overheen gaat, "
    "wordt de ophaling automatisch doorgeschoven naar de volgende ophaaldag."
)


_TABS = ["Overzicht", "Toevoegen", "Bewerken", "Importeren vanuit xlsx"]


def _nav(tab: str) -> None:
    st.session_state["_beheer_tab"] = tab


def render():
    st.title("Monsternemer beheer")
    st.caption("Deze data bevat persoonsgegevens. De database is lokaal en staat niet in de repo.")

    initialiseer_db()

    if "_beheer_tab" not in st.session_state:
        st.session_state["_beheer_tab"] = "Overzicht"

    actief = st.session_state["_beheer_tab"]

    cols = st.columns(len(_TABS))
    for col, label in zip(cols, _TABS, strict=False):
        with col:
            if label == actief:
                st.markdown(
                    f'<div style="border-bottom:2px solid #f63366;padding-bottom:6px;'
                    f'font-weight:600;color:#f63366">{label}</div>',
                    unsafe_allow_html=True,
                )
            else:
                st.button(label, on_click=_nav, args=(label,), use_container_width=True)

    st.divider()

    if actief == "Overzicht":
        _render_overzicht()
    elif actief == "Toevoegen":
        _render_toevoegen()
    elif actief == "Bewerken":
        _render_bewerken()
    elif actief == "Importeren vanuit xlsx":
        _render_import()


def _render_overzicht_tabel(monsternemers: list) -> None:
    st.caption(f"{len(monsternemers)} monsternemers in database")
    data = [
        {
            "ID": m.id,
            "Naam": m.volledige_naam,
            "Adres": m.adres or "",
            "Postcode": m.postcode,
            "Woonplaats": m.woonplaats,
            "Land": m.land or "",
            "Mobiel": m.telefoon or "",
            "Ophaaldagen": ", ".join(m.ophaaldagen),
            "Laadinstructie": m.laadinstructie or "",
            "Aantal lege bakken": m.aantal_lege_bakken,
            "Sjabloon": "ja" if m.sjabloon else "",
            "Uiterlijke tijd": m.uiterlijke_tijd or "",
            "Uiterlijke plantijd": m.uiterlijke_plantijd or "",
            "Bijzonderheden": m.bijzonderheden or "",
            "Ophalen": "ja" if m.ophalen else "nee (brengt zelf)",
        }
        for m in monsternemers
    ]
    st.dataframe(
        pd.DataFrame(data),
        use_container_width=True,
        hide_index=True,
        column_config={
            "Uiterlijke tijd": st.column_config.TextColumn(
                "Uiterlijke tijd", help=_UITERLIJKE_TIJD_HELP
            ),
            "Uiterlijke plantijd": st.column_config.TextColumn(
                "Uiterlijke plantijd", help=_UITERLIJKE_PLANTIJD_HELP
            ),
        },
    )


def _render_overzicht():
    if msg := st.session_state.pop("_beheer_success", None):
        st.success(msg)

    monsternemers = haal_alle_monsternemers()

    if not monsternemers:
        st.info("Geen monsternemers in de database. Importeer ze via het tabblad Importeren.")
        return

    _render_overzicht_tabel(monsternemers)

    st.divider()
    st.subheader("Monsternemer verwijderen")
    verwijder_id = st.number_input("ID van te verwijderen monsternemer", min_value=1, step=1)
    if st.button("Verwijder", type="secondary"):
        if verwijder_monsternemer(int(verwijder_id)):
            st.success(f"Monsternemer #{verwijder_id} verwijderd.")
            st.rerun()
        else:
            st.error(f"Geen monsternemer gevonden met ID {verwijder_id}.")


def _monsternemer_form(prefix: str, m: Monsternemer | None = None) -> dict:
    """Render het formulier en retourneer de ingevulde waarden als dict."""
    col1, col2, col3 = st.columns(3)
    with col1:
        voornaam = st.text_input(
            "Voornaam *", value=m.voornaam if m else "", key=f"{prefix}_voornaam"
        )
    with col2:
        tussenvoegsel = st.text_input(
            "Tussenvoegsel", value=m.tussenvoegsel or "" if m else "", key=f"{prefix}_tussenvoegsel"
        )
    with col3:
        achternaam = st.text_input(
            "Achternaam *", value=m.achternaam if m else "", key=f"{prefix}_achternaam"
        )

    col4, col5 = st.columns(2)
    with col4:
        adres = st.text_input(
            "Adres (straat + huisnummer)", value=m.adres or "" if m else "", key=f"{prefix}_adres"
        )
        postcode = st.text_input(
            "Postcode", value=m.postcode or "" if m else "", key=f"{prefix}_postcode"
        )
    with col5:
        woonplaats = st.text_input(
            "Woonplaats", value=m.woonplaats or "" if m else "", key=f"{prefix}_woonplaats"
        )
        land = st.text_input("Land", value=m.land or "" if m else "Nederland", key=f"{prefix}_land")
    telefoon = st.text_input(
        "Telefoonnummer", value=m.telefoon or "" if m else "", key=f"{prefix}_telefoon"
    )

    ophaaldagen_opties = ["ma", "di", "wo", "do", "vr", "za", "zo"]
    ophaaldagen = st.multiselect(
        "Ophaaldagen *",
        options=ophaaldagen_opties,
        default=[d for d in m.ophaaldagen if d in ophaaldagen_opties] if m else [],
        key=f"{prefix}_ophaaldagen",
    )

    laadinstructie = st.text_area(
        "Laadinstructie chauffeur",
        value=m.laadinstructie or "" if m else "",
        key=f"{prefix}_laadinstructie",
    )

    col_ut, col_upt = st.columns(2)
    with col_ut:
        uiterlijke_tijd = st.text_input(
            "Uiterlijke tijd (HH:MM)",
            value=m.uiterlijke_tijd or "" if m else "",
            placeholder="21:30",
            help=_UITERLIJKE_TIJD_HELP,
            key=f"{prefix}_uiterlijke_tijd",
        )
    with col_upt:
        uiterlijke_plantijd = st.text_input(
            "Uiterlijke plantijd (HH:MM)",
            value=m.uiterlijke_plantijd or "" if m else "",
            placeholder="21:30",
            help=_UITERLIJKE_PLANTIJD_HELP,
            key=f"{prefix}_uiterlijke_plantijd",
        )

    bijzonderheden = st.text_input(
        "Bijzonderheden (bijv. 'Sleutel nodig')",
        value=m.bijzonderheden or "" if m else "",
        key=f"{prefix}_bijzonderheden",
    )

    col6, col7 = st.columns(2)
    with col6:
        aantal_lege_bakken = st.number_input(
            "Aantal lege bakken meenemen",
            min_value=0,
            value=m.aantal_lege_bakken if m else 2,
            step=1,
            key=f"{prefix}_bakken",
        )
    with col7:
        sjabloon = st.checkbox(
            "Sjabloon aanwezig", value=m.sjabloon if m else False, key=f"{prefix}_sjabloon"
        )

    ophalen = st.checkbox("Wij halen op", value=m.ophalen if m else True, key=f"{prefix}_ophalen")

    return {
        "voornaam": voornaam,
        "tussenvoegsel": tussenvoegsel,
        "achternaam": achternaam,
        "adres": adres,
        "postcode": postcode,
        "woonplaats": woonplaats,
        "land": land,
        "telefoon": telefoon,
        "ophaaldagen": ophaaldagen,
        "laadinstructie": laadinstructie,
        "uiterlijke_tijd": uiterlijke_tijd,
        "uiterlijke_plantijd": uiterlijke_plantijd,
        "bijzonderheden": bijzonderheden,
        "aantal_lege_bakken": int(aantal_lege_bakken),
        "sjabloon": sjabloon,
        "ophalen": ophalen,
    }


def _dict_naar_monsternemer(data: dict, bestaande: Monsternemer | None = None) -> Monsternemer:
    return Monsternemer(
        id=bestaande.id if bestaande else None,
        code=bestaande.code if bestaande else "AP06",
        voornaam=data["voornaam"].strip(),
        tussenvoegsel=data["tussenvoegsel"].strip() or None,
        achternaam=data["achternaam"].strip(),
        adres=data["adres"].strip(),
        postcode=data["postcode"].strip().upper(),
        woonplaats=data["woonplaats"].strip(),
        land=data["land"].strip() or None,
        telefoon=data["telefoon"].strip() or None,
        laadinstructie=data["laadinstructie"].strip() or None,
        ophaaldagen=data["ophaaldagen"],
        uiterlijke_tijd=data["uiterlijke_tijd"].strip() or None,
        uiterlijke_plantijd=data["uiterlijke_plantijd"].strip() or None,
        bijzonderheden=data["bijzonderheden"].strip() or None,
        aantal_lege_bakken=data["aantal_lege_bakken"],
        sjabloon=data["sjabloon"],
        ophalen=data["ophalen"],
    )


def _render_toevoegen():
    st.subheader("Nieuwe monsternemer toevoegen")

    with st.form("toevoegen_form"):
        data = _monsternemer_form("add")
        submitted = st.form_submit_button("Toevoegen", type="primary")

    if submitted:
        ophaaldagen_ok = not data["ophalen"] or bool(data["ophaaldagen"])
        if not data["voornaam"] or not data["achternaam"] or not ophaaldagen_ok:
            st.error(
                "Vul minstens voornaam en achternaam in. Ophaaldagen zijn verplicht als wij ophalen."
            )
            return
        m = _dict_naar_monsternemer(data)
        nieuw_id = voeg_monsternemer_toe(m)
        st.session_state["_beheer_success"] = f"{m.volledige_naam} toegevoegd (ID: {nieuw_id})."
        st.rerun()


def _heeft_wijzigingen(nieuw: "Monsternemer", oud: "Monsternemer") -> bool:
    return (
        nieuw.voornaam != oud.voornaam
        or nieuw.tussenvoegsel != oud.tussenvoegsel
        or nieuw.achternaam != oud.achternaam
        or nieuw.adres != oud.adres
        or nieuw.postcode != oud.postcode
        or nieuw.woonplaats != oud.woonplaats
        or nieuw.land != oud.land
        or nieuw.telefoon != oud.telefoon
        or nieuw.ophaaldagen != oud.ophaaldagen
        or nieuw.laadinstructie != oud.laadinstructie
        or nieuw.uiterlijke_tijd != oud.uiterlijke_tijd
        or nieuw.uiterlijke_plantijd != oud.uiterlijke_plantijd
        or nieuw.bijzonderheden != oud.bijzonderheden
        or nieuw.aantal_lege_bakken != oud.aantal_lege_bakken
        or nieuw.sjabloon != oud.sjabloon
        or nieuw.ophalen != oud.ophalen
    )


def _render_bewerken():
    monsternemers = haal_alle_monsternemers()
    if not monsternemers:
        st.info("Geen monsternemers in de database.")
        return

    opties = {f"{m.id} — {m.volledige_naam}": m for m in monsternemers}

    if "_bewerken_form_uid" not in st.session_state:
        st.session_state["_bewerken_form_uid"] = uuid.uuid4().hex

    def _reset_form():
        st.session_state["_bewerken_form_uid"] = uuid.uuid4().hex

    keuze = st.selectbox(
        "Selecteer monsternemer",
        options=list(opties.keys()),
        on_change=_reset_form,
    )
    m = opties[keuze]
    uid = st.session_state["_bewerken_form_uid"]

    with st.form(f"bewerken_form_{uid}"):
        data = _monsternemer_form(uid, m)
        submitted = st.form_submit_button("Opslaan", type="primary")

    if submitted:
        ophaaldagen_ok = not data["ophalen"] or bool(data["ophaaldagen"])
        if not data["voornaam"] or not data["achternaam"] or not ophaaldagen_ok:
            st.error(
                "Vul minstens voornaam en achternaam in. Ophaaldagen zijn verplicht als wij ophalen."
            )
            return
        gewijzigd = _dict_naar_monsternemer(data, bestaande=m)
        if not _heeft_wijzigingen(gewijzigd, m):
            st.info("Geen wijzigingen — gegevens zijn al actueel.")
            return
        if update_monsternemer(gewijzigd):
            st.session_state["_beheer_success"] = f"✓ {gewijzigd.volledige_naam} opgeslagen."
            _nav("Overzicht")
            st.rerun()
        else:
            st.error("Opslaan mislukt.")


def _render_import():
    st.subheader("Importeren vanuit AP06 xlsx-bestand")
    st.caption(
        "Verwacht het standaard AP06 monsternemer-bestand met kolommen: "
        "Code, Voornaam, Tussenvoegsel, Achternaam, Adres, Postcode, Plaats, "
        "Land, Mobiel, Laadinstructie, Ophaaldagen, Sjabloon, Uiterlijke tijd, Bijzonderheden"
    )

    uploaded = st.file_uploader(
        "Kies het monsternemer-xlsx bestand",
        type=["xlsx"],
        key="import_upload",
    )

    if not uploaded:
        return

    bytes_data = BytesIO(uploaded.read())
    wb = load_workbook(bytes_data, data_only=True)
    ws = wb.active

    preview_data = []
    monsternemers_te_importeren = []

    for _i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if all(v is None for v in row):
            break
        if len(row) < 9:
            continue

        code = str(row[0]).strip() if row[0] else "AP06"
        voornaam = str(row[1]).strip() if row[1] else ""
        tussenvoegsel = str(row[2]).strip() if row[2] else None
        achternaam = str(row[3]).strip() if row[3] else ""
        adres = str(row[4]).strip() if row[4] else ""
        postcode = str(row[5]).strip() if row[5] else ""
        woonplaats = str(row[6]).strip() if row[6] else ""
        land = str(row[7]).strip() if (len(row) > 7 and row[7]) else None
        telefoon = str(row[8]).strip().replace("'", "") if row[8] else None
        laadinstructie = str(row[9]).strip() if (len(row) > 9 and row[9]) else None
        ophaaldagen_str = str(row[10]).strip() if (len(row) > 10 and row[10]) else ""
        ophaaldagen = [d.strip() for d in ophaaldagen_str.split(",") if d.strip()]
        aantal_lege_bakken = int(row[11]) if (len(row) > 11 and row[11] is not None) else 2
        sjabloon = bool(row[12]) if (len(row) > 12 and row[12] is not None) else False
        uiterlijke_tijd = None
        if len(row) > 13 and row[13]:
            ut = row[13]
            uiterlijke_tijd = ut.strftime("%H:%M") if hasattr(ut, "strftime") else str(ut).strip()
        bijzonderheden = str(row[14]).strip() if (len(row) > 14 and row[14]) else None

        if not voornaam or not achternaam:
            continue

        m = Monsternemer(
            id=None,
            code=code,
            voornaam=voornaam,
            tussenvoegsel=tussenvoegsel or None,
            achternaam=achternaam,
            adres=adres,
            postcode=postcode,
            woonplaats=woonplaats,
            land=land,
            telefoon=telefoon,
            laadinstructie=laadinstructie,
            ophaaldagen=ophaaldagen,
            uiterlijke_tijd=uiterlijke_tijd,
            uiterlijke_plantijd=None,
            bijzonderheden=bijzonderheden,
            aantal_lege_bakken=aantal_lege_bakken,
            sjabloon=sjabloon,
            ophalen=True,
        )
        monsternemers_te_importeren.append(m)
        preview_data.append(
            {
                "Naam": m.volledige_naam,
                "Woonplaats": woonplaats,
                "Ophaaldagen": ophaaldagen_str,
            }
        )

    wb.close()

    if not preview_data:
        st.warning("Geen geldige monsternemers gevonden in het bestand.")
        return

    st.info(f"Gevonden: {len(preview_data)} monsternemers")
    st.dataframe(pd.DataFrame(preview_data), use_container_width=True, hide_index=True)

    if st.button("Importeer alle monsternemers", type="primary"):
        for m in monsternemers_te_importeren:
            voeg_monsternemer_toe(m)
        st.success(f"{len(monsternemers_te_importeren)} monsternemers geimporteerd.")
        st.rerun()
